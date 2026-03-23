from pathlib import Path
import csv
from datetime import datetime, timezone

from db import get_connection


def _read_rows_csv(path: Path) -> list[dict[str, str]]:
    for encoding in ("utf-8-sig", "latin-1"):
        try:
            with path.open(newline="", encoding=encoding) as f:
                reader = csv.DictReader(f)
                if reader.fieldnames is None:
                    print(f"[WARN] CSV has no header row: {path}")
                    return []
                field_map = {name.strip().lower(): name for name in reader.fieldnames}
                rows = []
                for row in reader:
                    rows.append({k: row[v] for k, v in field_map.items()})
                return rows
        except UnicodeDecodeError:
            continue
    raise ValueError(f"Cannot decode CSV file {path} as UTF-8 or Latin-1")


def _read_rows_xlsx(path: Path) -> list[dict[str, str]]:
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)

    headers = next(rows_iter, None)
    if headers is None:
        wb.close()
        print(f"[WARN] XLSX file is empty (no header row): {path}")
        return []

    field_map = {str(h).strip().lower(): i for i, h in enumerate(headers) if h is not None}

    rows = []
    for row in rows_iter:
        rows.append({k: str(row[i]).strip() if row[i] is not None else "" for k, i in field_map.items()})
    wb.close()
    return rows


def load_po_master(file_path: Path) -> dict:
    conn = get_connection()
    cur = conn.cursor()

    now = datetime.now(timezone.utc).isoformat()
    inserted = 0

    ext = file_path.suffix.lower()
    if ext == ".xlsx":
        rows = _read_rows_xlsx(file_path)
    else:
        rows = _read_rows_csv(file_path)

    if not rows:
        conn.close()
        return {"po_loaded": 0, "source_file": str(file_path), "imported_at": now}

    sample_keys = set(rows[0].keys())
    po_key = next((k for k in sample_keys if k == "purchase order"), None)
    supplier_key = next((k for k in sample_keys if k == "supplier account"), None)
    status_key = next((k for k in sample_keys if k == "purchase order status"), None)
    approval_key = next((k for k in sample_keys if k == "approval status"), None)

    if not po_key or not supplier_key:
        conn.close()
        raise ValueError(f"Required columns not found. Got: {list(sample_keys)}")

    try:
        conn.execute("BEGIN")

        # Clear existing data (V1 replace semantics)
        cur.execute("DELETE FROM po_master")

        for row in rows:
            po_number = row.get(po_key, "").strip()
            supplier_account = row.get(supplier_key, "").strip()
            po_status = row.get(status_key, "").strip() if status_key else None
            approval_status = row.get(approval_key, "").strip() if approval_key else None

            if not po_number or not supplier_account:
                continue

            cur.execute(
                """
                INSERT INTO po_master (
                    po_number,
                    supplier_account,
                    po_status,
                    approval_status,
                    last_import_datetime
                ) VALUES (?, ?, ?, ?, ?)
                """,
                (po_number, supplier_account, po_status, approval_status, now),
            )
            inserted += cur.rowcount

        conn.commit()

    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()

    return {
        "po_loaded": inserted,
        "source_file": str(file_path),
        "imported_at": now,
    }


if __name__ == "__main__":
    summary = load_po_master(Path("data/Purchase_orders.csv"))
    print(summary)
